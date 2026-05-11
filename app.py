import os, io, base64, json, re
from flask import Flask, request, send_file, jsonify, make_response
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

app = Flask(__name__)
CORS(app)

DARK  = "1A1A2E"
GREEN = "1D9E75"
WHITE = "FFFFFF"
LGREY = "F7F7F7"

DATABASE_URL = os.environ.get('DATABASE_URL')
GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN')
GITHUB_REPO  = 'elmadix1/bglam-commandes'
GITHUB_BRANCH = 'main'
IMAGES_BASE_URL = 'https://elmadix1.github.io/bglam-commandes/images'

def get_db():
    import pg8000
    import urllib.parse
    url = urllib.parse.urlparse(DATABASE_URL)
    return pg8000.connect(
        user=url.username,
        password=url.password,
        host=url.hostname,
        port=url.port or 5432,
        database=url.path.lstrip('/')
    )

def init_db():
    if not DATABASE_URL:
        return
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS catalogues (
                supplier   VARCHAR(50) PRIMARY KEY,
                items      TEXT NOT NULL,
                updated_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        conn.commit()
        cur.close()
        conn.close()
        print("DB init OK")
    except Exception as e:
        print(f"DB init error: {e}")

init_db()

def upload_image_to_github(ref, b64_data):
    """Upload a base64 image to GitHub repo and return its URL."""
    if not GITHUB_TOKEN:
        return None
    try:
        import urllib.request, urllib.error
        # Strip data URI prefix if present
        if b64_data.startswith('data:'):
            b64_data = b64_data.split(',', 1)[1]
        
        # Determine filename - use ref as filename
        safe_ref = re.sub(r'[^a-zA-Z0-9_\-]', '_', ref)
        filename = f"{safe_ref}.jpg"
        path = f"images/{filename}"
        api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}"
        
        # Check if file already exists (to get SHA for update)
        sha = None
        try:
            req = urllib.request.Request(api_url, headers={
                'Authorization': f'token {GITHUB_TOKEN}',
                'Accept': 'application/vnd.github.v3+json'
            })
            with urllib.request.urlopen(req) as resp:
                existing = json.loads(resp.read())
                sha = existing.get('sha')
        except urllib.error.HTTPError as he:
            if he.code == 404:
                pass  # File doesn't exist yet — will create
            else:
                pass  # Other error — try anyway
        
        # Upload file
        payload = {
            'message': f'Add image {filename}',
            'content': b64_data,
            'branch': GITHUB_BRANCH
        }
        if sha:
            payload['sha'] = sha
            payload['message'] = f'Update image {filename}'
        
        data = json.dumps(payload).encode('utf-8')
        req = urllib.request.Request(api_url, data=data, method='PUT', headers={
            'Authorization': f'token {GITHUB_TOKEN}',
            'Accept': 'application/vnd.github.v3+json',
            'Content-Type': 'application/json'
        })
        with urllib.request.urlopen(req) as resp:
            result = json.loads(resp.read())
            return f"{IMAGES_BASE_URL}/{filename}"
    except urllib.error.HTTPError as he:
        if he.code == 409:
            # File already exists — that's fine, return the URL
            return f"{IMAGES_BASE_URL}/{filename}"
        print(f"GitHub upload error for {ref}: {he}")
        return None
    except Exception as e:
        print(f"GitHub upload error for {ref}: {e}")
        return None

def hex_fill(h): return PatternFill("solid", fgColor=h)
def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)
def base64_to_pil(b64):
    if b64.startswith("data:"): b64 = b64.split(",",1)[1]
    return PILImage.open(io.BytesIO(base64.b64decode(b64))).convert("RGBA")
def pil_to_png(img, size=(80,80)):
    img.thumbnail(size, PILImage.LANCZOS)
    buf = io.BytesIO(); img.save(buf, format="PNG"); buf.seek(0); return buf

@app.after_request
def cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

@app.route('/')
def index():
    return jsonify({"status": "BGlam Excel API", "version": "1.3", "db": bool(DATABASE_URL)})

@app.route('/catalogue/<supplier>/save', methods=['POST','OPTIONS'])
def save_catalogue(supplier):
    if request.method == 'OPTIONS': return make_response('', 204)
    try:
        items = request.get_json(force=True).get('items', [])
        # Strip base64 images before storing
        light = []
        for item in items:
            c = dict(item)
            if c.get('img','').startswith('data:'): c['img'] = ''
            light.append(c)
        items_json = json.dumps(light, ensure_ascii=False)
        if DATABASE_URL:
            conn = get_db(); cur = conn.cursor()
            cur.execute('''
                INSERT INTO catalogues (supplier, items, updated_at) VALUES (%s, %s, NOW())
                ON CONFLICT (supplier) DO UPDATE SET items=EXCLUDED.items, updated_at=NOW()
            ''', (supplier, items_json))
            conn.commit(); cur.close(); conn.close()
            return jsonify({'ok': True, 'count': len(light), 'storage': 'postgresql'})
        return jsonify({'ok': False, 'error': 'No DB'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/catalogue/<supplier>/load', methods=['GET','OPTIONS'])
def load_catalogue(supplier):
    if request.method == 'OPTIONS': return make_response('', 204)
    try:
        if DATABASE_URL:
            conn = get_db(); cur = conn.cursor()
            cur.execute('SELECT items, updated_at FROM catalogues WHERE supplier=%s', (supplier,))
            row = cur.fetchone(); cur.close(); conn.close()
            if row:
                items = json.loads(row[0])
                return jsonify({'ok': True, 'items': items, 'count': len(items), 'saved_at': str(row[1])})
        return jsonify({'ok': False, 'items': [], 'count': 0})
    except Exception as e:
        return jsonify({'ok': False, 'items': [], 'count': 0, 'error': str(e)})

@app.route('/catalogue/<supplier>/clear', methods=['POST','OPTIONS'])
def clear_catalogue(supplier):
    if request.method == 'OPTIONS': return make_response('', 204)
    try:
        if DATABASE_URL:
            conn = get_db(); cur = conn.cursor()
            cur.execute('DELETE FROM catalogues WHERE supplier=%s', (supplier,))
            conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/extract-images', methods=['POST','OPTIONS'])
def extract_images():
    if request.method == 'OPTIONS': return make_response('', 204)
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(request.files['file'].read()))
        ws = wb.active
        header_row, ref_col = 1, 1
        REF_KW = ('item no','item no.','item no:','ref','reference','sku','code')
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
            for cell in row:
                if cell.value and str(cell.value).strip().lower() in REF_KW:
                    header_row, ref_col = i, cell.column; break
            else: continue
            break
        row_to_ref = {}
        for row in ws.iter_rows(min_row=header_row+1, values_only=False):
            for cell in row:
                if cell.column == ref_col and cell.value:
                    val = str(cell.value).strip()
                    if val.lower() not in REF_KW and val:
                        row_to_ref[cell.row] = val; break
        img_map = {}
        for img in ws._images:
            try:
                anchor = img.anchor
                if hasattr(anchor, '_from'):
                    r = anchor._from.row + 1
                    d = img._data()
                    if d and len(d) > 100:
                        img_map[r] = 'data:image/png;base64,' + base64.b64encode(d).decode()
            except: pass
        result = {}
        for img_row, b64 in img_map.items():
            ref = row_to_ref.get(img_row)
            if not ref:
                for off in range(-5, 6):
                    ref = row_to_ref.get(img_row+off)
                    if ref: break
            if ref and ref not in result:
                result[ref] = b64

        # Upload to GitHub in background thread (non-blocking)
        if GITHUB_TOKEN and result:
            import threading
            def upload_all():
                for ref, b64 in result.items():
                    upload_image_to_github(ref, b64)
            t = threading.Thread(target=upload_all, daemon=True)
            t.start()

        # Build URL map based on what will be uploaded
        url_map = {}
        if GITHUB_TOKEN:
            for ref in result:
                safe_ref = re.sub(r'[^a-zA-Z0-9_\-]', '_', ref)
                url_map[ref] = f"{IMAGES_BASE_URL}/{safe_ref}.jpg"

        return jsonify({
            "images": result,
            "urls": url_map,
            "matched": len(result),
            "total_images": len(img_map),
            "total_refs": len(row_to_ref)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/excel', methods=['POST','OPTIONS'])
def generate_excel():
    if request.method == 'OPTIONS': return make_response('', 204)
    try: data = request.get_json(force=True)
    except: return jsonify({"error": "Invalid JSON"}), 400

    supplier   = data.get("supplier",   "BGlam")
    order_name = data.get("order_name", "Commande")
    date_str   = data.get("date",       "")
    items      = data.get("items",      [])
    if not items: return jsonify({"error": "No items"}), 400

    wb = Workbook(); ws = wb.active; ws.title = "Commande"
    ws.merge_cells("A1:K1"); ws["A1"] = f"{supplier} — Bon de Commande"
    ws["A1"].font = Font(name="Arial",size=14,bold=True,color=WHITE)
    ws["A1"].fill = hex_fill(DARK)
    ws["A1"].alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2"); ws["A2"] = f"Commande : {order_name}"
    ws["A2"].font = Font(name="Arial",size=11,bold=True); ws["A2"].fill = hex_fill(LGREY)
    ws["A2"].alignment = Alignment(vertical="center")
    ws.merge_cells("E2:H2"); ws["E2"] = f"Date : {date_str}"
    ws["E2"].font = Font(name="Arial",size=11); ws["E2"].fill = hex_fill(LGREY)
    ws["E2"].alignment = Alignment(vertical="center"); ws.row_dimensions[2].height = 22

    headers=["Photo","CTN NO","ITEM NO","DESCRIPTION","CATEGORIE","PRICE","PCS/CTN","CTN","QTY","AMOUNT","REMARK"]
    widths =[14,      10,      12,       28,            22,         10,     10,       6,    8,    12,      35]
    HR=3
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(row=HR,column=ci,value=h)
        cell.font=Font(name="Arial",size=10,bold=True,color=WHITE)
        cell.fill=hex_fill(DARK); cell.alignment=Alignment(horizontal="center",vertical="center")
        cell.border=thin_border()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[HR].height=20

    total=0.0
    for ri,item in enumerate(items):
        r=HR+1+ri; qty=item.get("qty") or 0; price=item.get("price")
        amt=round(qty*price,2) if price is not None else None
        if amt: total+=amt
        bg=LGREY if ri%2==0 else WHITE
        vals=[item.get("group",""),item.get("ref",""),item.get("desc",""),
              item.get("cat",""),price,item.get("moq"),"",qty,amt,item.get("remark","")]
        for ci,val in enumerate(vals,2):
            cell=ws.cell(row=r,column=ci,value=val)
            cell.font=Font(name="Arial",size=10); cell.fill=hex_fill(bg)
            cell.border=thin_border(); cell.alignment=Alignment(vertical="center")
            if ci in (6,7,9,10): cell.alignment=Alignment(horizontal="right",vertical="center")
            if ci==10 and amt is not None: cell.font=Font(name="Arial",size=10,bold=True,color=GREEN)
        ws.row_dimensions[r].height=62
        img_b64=item.get("img","")
        if img_b64 and len(img_b64)>100:
            try:
                pil=base64_to_pil(img_b64); buf=pil_to_png(pil,(60,60))
                xl=XLImage(buf); xl.anchor=f"A{r}"; ws.add_image(xl)
            except: pass
        cell=ws.cell(row=r,column=1,value=""); cell.fill=hex_fill(bg); cell.border=thin_border()

    tr=HR+1+len(items); tq=sum((item.get("qty") or 0) for item in items)
    ws.merge_cells(f"A{tr}:H{tr}")
    c=ws.cell(row=tr,column=1,value="TOTAL")
    c.font=Font(name="Arial",size=11,bold=True,color=WHITE); c.fill=hex_fill(DARK)
    c.alignment=Alignment(horizontal="right",vertical="center")
    c=ws.cell(row=tr,column=9,value=tq)
    c.font=Font(name="Arial",size=11,bold=True,color=WHITE); c.fill=hex_fill(DARK)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c=ws.cell(row=tr,column=10,value=round(total,2))
    c.font=Font(name="Arial",size=12,bold=True,color="4ADE80"); c.fill=hex_fill(DARK)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[tr].height=24; ws.freeze_panes=f"B{HR+1}"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    safe=re.sub(r"[^a-zA-Z0-9_\- ]","_",f"{supplier}_{order_name}")
    fname=f"{safe}_{date_str.replace('//','-')}.xlsx"
    resp=make_response(send_file(buf,as_attachment=True,download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
    resp.headers['Access-Control-Allow-Origin']='*'
    return resp

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
